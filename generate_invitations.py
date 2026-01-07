#!/usr/bin/env python3
import argparse
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw, ImageFont

IMAGE_WIDTH = 680
IMAGE_HEIGHT = 960
EMBED_HEIGHT = 540  # px; keeps Excel row height <= 409 points
EMBED_WIDTH = int(IMAGE_WIDTH * (EMBED_HEIGHT / IMAGE_HEIGHT))

TEXT_COLOR = (122, 111, 93)  # #7a6f5d
TEXT_DARK = (90, 79, 61)  # #5a4f3d
ACCENT_COLOR = (139, 157, 195)  # #8b9dc3
DIVIDER_COLOR = (196, 181, 160)  # #c4b5a0


def normalize_header(value):
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text).strip().lower()


def find_column(headers, targets):
    targets = {normalize_header(t) for t in targets}
    for idx, header in enumerate(headers, start=1):
        if normalize_header(header) in targets:
            return idx
    return None


def first_empty_header_column(headers):
    for idx, header in enumerate(headers, start=1):
        if header is None or str(header).strip() == "":
            return idx
    return len(headers) + 1


def safe_slug(text):
    text = normalize_header(text)
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "guest"


def font_from_paths(paths, size):
    for path in paths:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def text_size(draw, text, font):
    left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
    return right - left, bottom - top


def draw_centered_text(draw, text, font, y, color, image_width, max_width=None, line_spacing=6):
    lines = wrap_text(draw, text, font, max_width or image_width)
    for line in lines:
        width, height = text_size(draw, line, font)
        x = (image_width - width) / 2
        draw.text((x, y), line, font=font, fill=color)
        y += height + line_spacing
    return y - line_spacing


def draw_centered_at(draw, text, font, center_x, y, color):
    width, _ = text_size(draw, text, font)
    draw.text((center_x - width / 2, y), text, font=font, fill=color)


def wrap_text(draw, text, font, max_width):
    if not text:
        return [""]
    words = text.split()
    if not words:
        return [text]
    lines = []
    current = words[0]
    for word in words[1:]:
        test = f"{current} {word}"
        if text_size(draw, test, font)[0] <= max_width:
            current = test
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def create_invitation_image(guest_name, output_path, monogram_path, fonts):
    image = Image.new("RGB", (IMAGE_WIDTH, IMAGE_HEIGHT), "white")
    draw = ImageDraw.Draw(image)

    y = 40

    if monogram_path and Path(monogram_path).exists():
        monogram = Image.open(monogram_path).convert("RGBA")
        target_width = 170
        scale = target_width / monogram.width
        target_height = int(monogram.height * scale)
        monogram = monogram.resize((target_width, target_height), Image.LANCZOS)
        x = int((IMAGE_WIDTH - target_width) / 2)
        image.paste(monogram, (x, y), monogram)
        y += target_height + 20
    else:
        y += 20

    y = draw_centered_text(draw, "TRÂN TRỌNG KÍNH MỜI", fonts["header"], y, TEXT_COLOR, IMAGE_WIDTH)
    y += 20
    y = draw_centered_text(
        draw,
        guest_name,
        fonts["guest"],
        y,
        TEXT_DARK,
        IMAGE_WIDTH,
        max_width=IMAGE_WIDTH - 120,
        line_spacing=8,
    )
    y += 25

    y += 15
    divider_width = 500
    draw.line(
        (
            (IMAGE_WIDTH - divider_width) / 2,
            y,
            (IMAGE_WIDTH + divider_width) / 2,
            y,
        ),
        fill=DIVIDER_COLOR,
        width=1,
    )
    y += 25

    y = draw_centered_text(draw, "TỚI DỰ BỮA CƠM THÂN MẬT", fonts["small"], y, TEXT_COLOR, IMAGE_WIDTH)
    y += 10
    y = draw_centered_text(
        draw,
        "MỪNG LỄ VU QUY CỦA HAI CON CHÚNG TÔI",
        fonts["small"],
        y,
        TEXT_COLOR,
        IMAGE_WIDTH,
    )
    y += 28

    y = draw_centered_text(draw, "Nguyễn Thị Hải Anh", fonts["couple"], y, ACCENT_COLOR, IMAGE_WIDTH)
    y += 6
    y = draw_centered_text(draw, "&", fonts["amp"], y, ACCENT_COLOR, IMAGE_WIDTH)
    y += 6
    y = draw_centered_text(draw, "Tạ Quang Thái", fonts["couple"], y, ACCENT_COLOR, IMAGE_WIDTH)
    y += 26

    y = draw_centered_text(draw, "VÀO HỒI 11 GIỜ 00, THỨ BA", fonts["small"], y, TEXT_COLOR, IMAGE_WIDTH)
    y += 12

    date_items = [("NGÀY", "20"), ("THÁNG", "01"), ("NĂM", "2026")]
    column_width = 100
    gap = 48
    total_width = column_width * 3 + gap * 2
    start_x = (IMAGE_WIDTH - total_width) / 2
    label_height = text_size(draw, "NGÀY", fonts["label"])[1]
    value_height = text_size(draw, "20", fonts["date"])[1]
    for idx, (label, value) in enumerate(date_items):
        center_x = start_x + column_width / 2 + idx * (column_width + gap)
        draw_centered_at(draw, label, fonts["label"], center_x, y, TEXT_COLOR)
        draw_centered_at(draw, value, fonts["date"], center_x, y + label_height + 8, TEXT_DARK)

    y = y + label_height + 8 + value_height + 12
    y = draw_centered_text(
        draw,
        "TỨC NGÀY 02 THÁNG 12 NĂM ẤT TỴ",
        fonts["small"],
        y,
        TEXT_COLOR,
        IMAGE_WIDTH,
    )
    y += 24

    y = draw_centered_text(draw, "TẠI TẦNG 2", fonts["small"], y, TEXT_COLOR, IMAGE_WIDTH)
    y += 8
    y = draw_centered_text(draw, "LONG VĨ PALACE", fonts["venue"], y, TEXT_DARK, IMAGE_WIDTH)
    y += 8
    y = draw_centered_text(
        draw,
        "SỐ 3A ĐẠO DUY ANH, ĐỐNG ĐA, HÀ NỘI",
        fonts["small"],
        y,
        TEXT_COLOR,
        IMAGE_WIDTH,
    )
    y += 24

    left_center = IMAGE_WIDTH * 0.25
    right_center = IMAGE_WIDTH * 0.75
    draw_centered_at(draw, "NHÀ GÁI", fonts["label"], left_center, y, TEXT_COLOR)
    draw_centered_at(draw, "NHÀ TRAI", fonts["label"], right_center, y, TEXT_COLOR)

    name_height = text_size(draw, "ÔNG NGUYỄN VĂN LUÂN", fonts["small"])[1]
    y_names = y + label_height + 8
    draw_centered_at(draw, "ÔNG NGUYỄN VĂN LUÂN", fonts["small"], left_center, y_names, TEXT_COLOR)
    draw_centered_at(draw, "ÔNG TẠ QUANG SÁNG", fonts["small"], right_center, y_names, TEXT_COLOR)
    y_names_2 = y_names + name_height + 4
    draw_centered_at(draw, "BÀ NGUYỄN THỊ MAI HƯƠNG", fonts["small"], left_center, y_names_2, TEXT_COLOR)
    draw_centered_at(draw, "BÀ NGUYỄN PHƯƠNG DUNG", fonts["small"], right_center, y_names_2, TEXT_COLOR)

    y = y_names_2 + name_height + 26
    draw_centered_text(draw, "RẤT HÂN HẠNH ĐƯỢC ĐÓN TIẾP!", fonts["small"], y, TEXT_COLOR, IMAGE_WIDTH)

    image.save(output_path, "PNG")


def excel_column_width(pixels):
    return max(10, round((pixels - 5) / 7, 2))


def main():
    parser = argparse.ArgumentParser(description="Generate wedding invitation images and embed into Excel.")
    parser.add_argument(
        "--input",
        default="Danh sach moi cuoi AA CH TM - chi Hop da sua.xlsx",
        help="Input Excel file path.",
    )
    parser.add_argument(
        "--output",
        default="Danh sach moi cuoi AA CH TM - chi Hop da sua_with_invitations.xlsx",
        help="Output Excel file path.",
    )
    parser.add_argument(
        "--output-dir",
        default="generated_invitations",
        help="Directory to store generated invitation images.",
    )
    parser.add_argument(
        "--monogram",
        default="IMG_1984.JPG",
        help="Path to monogram image used in invitations.",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    fonts = {
        "header": font_from_paths(
            ["/System/Library/Fonts/Supplemental/Times New Roman.ttf"], 15
        ),
        "guest": font_from_paths(
            ["/System/Library/Fonts/Supplemental/Times New Roman Italic.ttf"], 28
        ),
        "small": font_from_paths(
            ["/System/Library/Fonts/Supplemental/Times New Roman.ttf"], 11
        ),
        "couple": font_from_paths(
            ["/System/Library/Fonts/Supplemental/Times New Roman Italic.ttf"], 42
        ),
        "amp": font_from_paths(
            ["/System/Library/Fonts/Supplemental/Times New Roman Italic.ttf"], 32
        ),
        "label": font_from_paths(
            ["/System/Library/Fonts/Supplemental/Times New Roman.ttf"], 10
        ),
        "date": font_from_paths(
            ["/System/Library/Fonts/Supplemental/Times New Roman.ttf"], 60
        ),
        "venue": font_from_paths(
            ["/System/Library/Fonts/Supplemental/Times New Roman Bold.ttf"], 18
        ),
    }

    workbook = load_workbook(args.input)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    name_col = find_column(headers, ["Kính gửi", "Kinh Gui"])
    if name_col is None:
        raise SystemExit("Could not find 'Kính gửi' column in the header row.")

    image_col = first_empty_header_column(headers)
    sheet.cell(row=1, column=image_col, value="Invitation Image")
    image_col_letter = sheet.cell(row=1, column=image_col).column_letter
    sheet.column_dimensions[image_col_letter].width = excel_column_width(EMBED_WIDTH)

    row_height = EMBED_HEIGHT * 0.75

    generated = 0
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row, column=name_col).value
        if value is None or str(value).strip() == "":
            continue

        guest_name = str(value).strip()
        slug = safe_slug(guest_name)
        filename = f"invitation_{row}_{slug}.png"
        output_path = output_dir / filename

        create_invitation_image(guest_name, output_path, args.monogram, fonts)

        xl_image = XLImage(str(output_path))
        xl_image.width = EMBED_WIDTH
        xl_image.height = EMBED_HEIGHT
        xl_image.anchor = sheet.cell(row=row, column=image_col).coordinate
        sheet.add_image(xl_image)

        sheet.row_dimensions[row].height = row_height
        generated += 1

        if generated % 50 == 0:
            print(f"Generated {generated} invitations...")

    workbook.save(args.output)
    print(f"Done. Generated {generated} invitations.")
    print(f"Images saved to: {output_dir}")
    print(f"Updated workbook: {args.output}")


if __name__ == "__main__":
    main()
